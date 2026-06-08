"""
AUTOMATIZADOR DE CRONOGRAMAS BIOMÉDICOS V2
==========================================
Con colores automáticos en Excel
🔴 ESTE MES | 🟡 PRÓXIMO | 🟠 EN 2 MESES | 🟢 OK
"""

import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

# ============================================================
# CONFIGURACIÓN — Cambia el nombre si tu archivo es diferente
# ============================================================
ARCHIVO_ENTRADA = "Cronograma_de_mantenimiento_ESE_Hospital_Yolombo.xlsx"
ARCHIVO_SALIDA  = "Reporte_Mantenimiento_V2.xlsx"

HOY = datetime.today()
MES_ACTUAL = HOY.month

MESES = {
    "ENERO":1,"FEBRERO":2,"MARZO":3,"ABRIL":4,"MAYO":5,"JUNIO":6,
    "JULIO":7,"AGOSTO":8,"SEPTIEMBRE":9,"OCTUBRE":10,"NOVIEMBRE":11,"DICIEMBRE":12
}
MESES_INVERSO = {v: k for k, v in MESES.items()}

COLORES = {
    "🔴 ESTE MES":    "FF4444",
    "🔴 VENCIDO":     "FF4444",
    "🟡 PRÓXIMO MES": "FFD700",
    "🟠 EN 2 MESES":  "FFA500",
    "🟢 OK":          "44BB44",
    "SIN FECHA":      "CCCCCC",
}
COLOR_ENCABEZADO = "1F3864"
COLOR_FILA_PAR   = "F2F2F2"

# ============================================================
# PROCESAR DATOS
# ============================================================
print("="*60)
print("AUTOMATIZADOR DE CRONOGRAMAS BIOMÉDICOS V2")
print("="*60)
print(f"Fecha: {HOY.strftime('%d/%m/%Y')}\n")

df = pd.read_excel(ARCHIVO_ENTRADA, header=0)
df.columns = ["ITEM","DESCRIPCION","UBICACION","N_INVENTARIO","FABRICANTE","MARCA",
              "MODELO","SERIE","CLASIFICACION_RIESGO","REGISTRO_INVIMA",
              "FECHA_COMPRA","PERIODICIDAD","FECHA_MANTENIMIENTO"]
df = df.dropna(subset=["DESCRIPCION"])
df.replace("NR","",inplace=True)

def extraer_meses(texto):
    if pd.isna(texto) or str(texto).strip()=="": return []
    return [MESES[p.strip()] for p in str(texto).upper().split("/") if p.strip() in MESES]

def calcular_estado(meses):
    if not meses: return "SIN FECHA", None
    meses = sorted(meses)
    proximo = next((m for m in meses if m >= MES_ACTUAL), None)
    if proximo is None: return "🔴 VENCIDO", meses[-1]
    diff = proximo - MES_ACTUAL
    if diff == 0: return "🔴 ESTE MES", proximo
    elif diff == 1: return "🟡 PRÓXIMO MES", proximo
    elif diff <= 2: return "🟠 EN 2 MESES", proximo
    else: return "🟢 OK", proximo

df["MESES_PARSED"] = df["FECHA_MANTENIMIENTO"].apply(extraer_meses)
resultados = df["MESES_PARSED"].apply(calcular_estado)
df["ESTADO"] = resultados.apply(lambda x: x[0])
df["PROXIMO_MES"] = resultados.apply(lambda x: MESES_INVERSO.get(x[1],"") if x[1] else "")

columnas = ["ITEM","DESCRIPCION","UBICACION","N_INVENTARIO",
            "PERIODICIDAD","FECHA_MANTENIMIENTO","PROXIMO_MES","ESTADO"]
df[columnas].to_excel(ARCHIVO_SALIDA, index=False, sheet_name="CRONOGRAMA")

# ============================================================
# APLICAR COLORES
# ============================================================
print("Aplicando colores...")

wb = load_workbook(ARCHIVO_SALIDA)
ws = wb["CRONOGRAMA"]

# Ancho columnas
for i, ancho in enumerate([6,35,25,15,12,22,15,18], 1):
    ws.column_dimensions[ws.cell(1,i).column_letter].width = ancho

# Encabezados
for cell in ws[1]:
    cell.fill = PatternFill("solid", fgColor=COLOR_ENCABEZADO)
    cell.font = Font(bold=True, color="FFFFFF", size=11)
    cell.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[1].height = 30

borde = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin")
)

# Filas con colores
for idx, fila in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=2):
    estado = ws.cell(idx, 8).value
    color_fila = COLOR_FILA_PAR if idx % 2 == 0 else "FFFFFF"

    for cell in fila[:-1]:
        cell.fill = PatternFill("solid", fgColor=color_fila)
        cell.alignment = Alignment(vertical="center")
        cell.border = borde

    # Celda ESTADO con color
    color_estado = COLORES.get(estado, "FFFFFF")
    fila[-1].fill = PatternFill("solid", fgColor=color_estado)
    fila[-1].font = Font(bold=True, color="FFFFFF" if estado in ["🔴 ESTE MES","🔴 VENCIDO","🟢 OK"] else "000000")
    fila[-1].alignment = Alignment(horizontal="center", vertical="center")
    fila[-1].border = borde

ws.freeze_panes = "A2"

# ============================================================
# HOJA RESUMEN
# ============================================================
ws_r = wb.create_sheet("RESUMEN")
datos = [
    ["RESUMEN DE MANTENIMIENTO", ""],
    [f"Fecha: {HOY.strftime('%d/%m/%Y')}", ""],
    ["",""],
    ["ESTADO", "CANTIDAD"],
    ["🔴 ESTE MES", int((df["ESTADO"]=="🔴 ESTE MES").sum())],
    ["🟡 PRÓXIMO MES", int((df["ESTADO"]=="🟡 PRÓXIMO MES").sum())],
    ["🟠 EN 2 MESES", int((df["ESTADO"]=="🟠 EN 2 MESES").sum())],
    ["🟢 OK", int((df["ESTADO"]=="🟢 OK").sum())],
    ["SIN FECHA", int((df["ESTADO"]=="SIN FECHA").sum())],
    ["",""],
    ["TOTAL EQUIPOS", len(df)],
]
for fila in datos:
    ws_r.append(fila)

ws_r.column_dimensions["A"].width = 25
ws_r.column_dimensions["B"].width = 15
ws_r["A1"].font = Font(bold=True, size=14, color=COLOR_ENCABEZADO)

fills = {4:COLOR_ENCABEZADO, 5:"FF4444", 6:"FFD700", 7:"FFA500", 8:"44BB44", 9:"CCCCCC"}
for row, color in fills.items():
    for col in ["A","B"]:
        c = ws_r[f"{col}{row}"]
        c.fill = PatternFill("solid", fgColor=color)
        c.font = Font(bold=True, color="FFFFFF" if row in [4,5,8] else "000000")
        c.alignment = Alignment(horizontal="center")

wb.save(ARCHIVO_SALIDA)

print(f"\n✅ Reporte V2 guardado: {ARCHIVO_SALIDA}")
print("\nRESUMEN:")
print(df["ESTADO"].value_counts().to_string())
print(f"\nTotal equipos: {len(df)}")
