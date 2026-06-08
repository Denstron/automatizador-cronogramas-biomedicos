"""
AUTOMATIZADOR DE CRONOGRAMAS BIOMÉDICOS V4
==========================================
- Todo lo de V3 +
- Envía email automático con equipos del mes
- Adjunta reporte PDF y Excel
"""

import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
import os

# ============================================================
# CONFIGURACIÓN — Edita estos datos
# ============================================================
ARCHIVO_ENTRADA = "Cronograma_de_mantenimiento_ESE_Hospital_Yolombo.xlsx"
ARCHIVO_EXCEL   = "Reporte_Cronograma_Equipos_Biomedicos.xlsx"
ARCHIVO_PDF     = "Reporte_Cronograma_Equipos_Biomedicos.pdf"

# Datos del remitente
EMAIL_REMITENTE  = "ingbiomedico.espana@gmail.com"
PASSWORD_APP     = "vkwp uqhu dhoe cjgk"  # Contraseña de aplicación

# Destinatarios (por ahora solo tú, luego agregas técnicos)
DESTINATARIOS = [
    {
        "nombre": "Daniel España",
        "email": "ingbiomedico.espana@gmail.com",
        "areas": []  # Lista vacía = recibe TODOS los equipos
    }
    # Cuando lleguen los técnicos, agrega así:
   #  {
     #    "nombre": "Maria Isabel",
      #   "email": "Lopezceballosmi@gmail.com",
       #  "areas": []
    #   },
]

HOY = datetime.today()
MES_ACTUAL = HOY.month
NOMBRE_MES = {1:"Enero",2:"Febrero",3:"Marzo",4:"Abril",5:"Mayo",6:"Junio",
              7:"Julio",8:"Agosto",9:"Septiembre",10:"Octubre",11:"Noviembre",12:"Diciembre"}

MESES = {"ENERO":1,"FEBRERO":2,"MARZO":3,"ABRIL":4,"MAYO":5,"JUNIO":6,
    "JULIO":7,"AGOSTO":8,"SEPTIEMBRE":9,"OCTUBRE":10,"NOVIEMBRE":11,"DICIEMBRE":12}
MESES_INVERSO = {v: k for k, v in MESES.items()}

COLORES_EXCEL = {"🔴 ESTE MES":"FF4444","🔴 VENCIDO":"FF4444",
    "🟡 PRÓXIMO MES":"FFD700","🟠 EN 2 MESES":"FFA500","🟢 OK":"44BB44","SIN FECHA":"CCCCCC"}
COLOR_ENCABEZADO = "1F3864"
COLOR_FILA_PAR   = "F2F2F2"

# ============================================================
# PROCESAR DATOS
# ============================================================
print("="*60)
print("AUTOMATIZADOR DE CRONOGRAMAS BIOMÉDICOS V4")
print("="*60)
print(f"Fecha: {HOY.strftime('%d/%m/%Y')}\n")

df = pd.read_excel(ARCHIVO_ENTRADA, header=0)
df.columns = ["ITEM","DESCRIPCION","UBICACION","N_INVENTARIO","FABRICANTE","MARCA",
              "MODELO","SERIE","CLASIFICACION_RIESGO","REGISTRO_INVIMA",
              "FECHA_COMPRA","PERIODICIDAD","FECHA_MANTENIMIENTO"]
df = df.dropna(subset=["DESCRIPCION"])
df.replace("NR","",inplace=True)

def extraer_meses(texto):
    if pd.isna(texto) or str(texto).strip()=="": return []
    return [MESES[p.strip()] for p in str(texto).upper().split("/") if p.strip() in MESES]

def calcular_estado(meses):
    if not meses: return "SIN FECHA", None
    meses = sorted(meses)
    proximo = next((m for m in meses if m >= MES_ACTUAL), None)
    if proximo is None: return "🔴 VENCIDO", meses[-1]
    diff = proximo - MES_ACTUAL
    if diff == 0: return "🔴 ESTE MES", proximo
    elif diff == 1: return "🟡 PRÓXIMO MES", proximo
    elif diff <= 2: return "🟠 EN 2 MESES", proximo
    else: return "🟢 OK", proximo

df["MESES_PARSED"] = df["FECHA_MANTENIMIENTO"].apply(extraer_meses)
resultados = df["MESES_PARSED"].apply(calcular_estado)
df["ESTADO"] = resultados.apply(lambda x: x[0])
df["PROXIMO_MES"] = resultados.apply(lambda x: MESES_INVERSO.get(x[1],"") if x[1] else "")

areas = sorted(df["UBICACION"].dropna().unique().tolist())
equipos_este_mes = df[df["ESTADO"] == "🔴 ESTE MES"]

print(f"Total equipos: {len(df)}")
print(f"Equipos este mes: {len(equipos_este_mes)}")

# ============================================================
# GENERAR EXCEL (igual que V3)
# ============================================================
print(f"\nGenerando Excel...")

columnas = ["ITEM","DESCRIPCION","UBICACION","N_INVENTARIO",
            "PERIODICIDAD","FECHA_MANTENIMIENTO","PROXIMO_MES","ESTADO"]
df_reporte = df[columnas].copy()

with pd.ExcelWriter(ARCHIVO_EXCEL, engine="openpyxl") as writer:
    df_reporte.to_excel(writer, sheet_name="TODOS", index=False)
    for area in areas:
        df_area = df_reporte[df_reporte["UBICACION"] == area]
        df_area.to_excel(writer, sheet_name=area[:31], index=False)

wb = load_workbook(ARCHIVO_EXCEL)

def aplicar_estilos(ws):
    for i, ancho in enumerate([6,35,25,15,12,22,15,18], 1):
        ws.column_dimensions[ws.cell(1,i).column_letter].width = ancho
    for cell in ws[1]:
        cell.fill = PatternFill("solid", fgColor=COLOR_ENCABEZADO)
        cell.font = Font(bold=True, color="FFFFFF", size=10)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 25
    borde = Border(left=Side(style="thin"),right=Side(style="thin"),
                   top=Side(style="thin"),bottom=Side(style="thin"))
    for idx, fila in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=2):
        estado = ws.cell(idx, 8).value
        color_fila = COLOR_FILA_PAR if idx % 2 == 0 else "FFFFFF"
        for cell in fila[:-1]:
            cell.fill = PatternFill("solid", fgColor=color_fila)
            cell.alignment = Alignment(vertical="center")
            cell.border = borde
        fila[-1].fill = PatternFill("solid", fgColor=COLORES_EXCEL.get(estado,"FFFFFF"))
        fila[-1].font = Font(bold=True, color="FFFFFF" if estado in ["🔴 ESTE MES","🔴 VENCIDO","🟢 OK"] else "000000")
        fila[-1].alignment = Alignment(horizontal="center", vertical="center")
        fila[-1].border = borde
    ws.freeze_panes = "A2"

for hoja in wb.sheetnames:
    aplicar_estilos(wb[hoja])

ws_r = wb.create_sheet("RESUMEN", 0)
for col, ancho in zip(["A","B","C","D","E","F"],[30,12,12,12,12,12]):
    ws_r.column_dimensions[col].width = ancho
ws_r.append(["ÁREA","🔴 ESTE MES","🟡 PRÓXIMO","🟠 2 MESES","🟢 OK","TOTAL"])
for cell in ws_r[1]:
    cell.fill = PatternFill("solid", fgColor=COLOR_ENCABEZADO)
    cell.font = Font(bold=True, color="FFFFFF")
    cell.alignment = Alignment(horizontal="center")
for area in areas:
    df_a = df[df["UBICACION"]==area]
    ws_r.append([area,
        int((df_a["ESTADO"]=="🔴 ESTE MES").sum()),
        int((df_a["ESTADO"]=="🟡 PRÓXIMO MES").sum()),
        int((df_a["ESTADO"]=="🟠 EN 2 MESES").sum()),
        int((df_a["ESTADO"]=="🟢 OK").sum()),
        len(df_a)])
ws_r.append(["TOTAL HOSPITAL",
    int((df["ESTADO"]=="🔴 ESTE MES").sum()),
    int((df["ESTADO"]=="🟡 PRÓXIMO MES").sum()),
    int((df["ESTADO"]=="🟠 EN 2 MESES").sum()),
    int((df["ESTADO"]=="🟢 OK").sum()),
    len(df)])
for cell in ws_r[ws_r.max_row]:
    cell.font = Font(bold=True)
    cell.fill = PatternFill("solid", fgColor="DDDDDD")
wb.save(ARCHIVO_EXCEL)
print("✅ Excel generado")

# ============================================================
# GENERAR PDF
# ============================================================
print("Generando PDF...")

doc = SimpleDocTemplate(ARCHIVO_PDF, pagesize=A4,
    topMargin=2*cm, bottomMargin=2*cm, leftMargin=2*cm, rightMargin=2*cm)
styles = getSampleStyleSheet()
titulo_style = ParagraphStyle("t", parent=styles["Title"], fontSize=16,
    textColor=colors.HexColor("#1F3864"), spaceAfter=8)
sub_style = ParagraphStyle("s", parent=styles["Normal"], fontSize=10,
    textColor=colors.grey, spaceAfter=20)

contenido = []
contenido.append(Paragraph("Reporte de Cronograma de Equipos Biomedicos", titulo_style))
contenido.append(Paragraph(f"Fecha: {HOY.strftime('%d/%m/%Y')} | Total equipos: {len(df)}", sub_style))
contenido.append(Paragraph("Resumen General", styles["Heading2"]))
contenido.append(Spacer(1, 0.3*cm))

datos_resumen = [
    ["Estado","Equipos","%"],
    ["Este mes", str(len(equipos_este_mes)), f"{len(equipos_este_mes)/len(df)*100:.1f}%"],
    ["Proximo mes", str(int((df["ESTADO"]=="🟡 PRÓXIMO MES").sum())), f"{int((df['ESTADO']=='🟡 PRÓXIMO MES').sum())/len(df)*100:.1f}%"],
    ["En 2 meses", str(int((df["ESTADO"]=="🟠 EN 2 MESES").sum())), f"{int((df['ESTADO']=='🟠 EN 2 MESES').sum())/len(df)*100:.1f}%"],
    ["OK", str(int((df["ESTADO"]=="🟢 OK").sum())), f"{int((df['ESTADO']=='🟢 OK').sum())/len(df)*100:.1f}%"],
    ["TOTAL", str(len(df)), "100%"],
]
t1 = Table(datos_resumen, colWidths=[8*cm,4*cm,4*cm])
t1.setStyle(TableStyle([
    ("BACKGROUND",(0,0),(-1,0),colors.HexColor("#1F3864")),
    ("TEXTCOLOR",(0,0),(-1,0),colors.white),
    ("FONTNAME",(0,0),(-1,0),"Helvetica-Bold"),
    ("ALIGN",(0,0),(-1,-1),"CENTER"),
    ("BACKGROUND",(0,1),(-1,1),colors.HexColor("#FF4444")),
    ("TEXTCOLOR",(0,1),(-1,1),colors.white),
    ("BACKGROUND",(0,2),(-1,2),colors.HexColor("#FFD700")),
    ("BACKGROUND",(0,3),(-1,3),colors.HexColor("#FFA500")),
    ("BACKGROUND",(0,4),(-1,4),colors.HexColor("#44BB44")),
    ("TEXTCOLOR",(0,4),(-1,4),colors.white),
    ("FONTNAME",(0,5),(-1,5),"Helvetica-Bold"),
    ("BACKGROUND",(0,5),(-1,5),colors.HexColor("#DDDDDD")),
    ("GRID",(0,0),(-1,-1),0.5,colors.grey),
]))
contenido.append(t1)
contenido.append(Spacer(1, 0.8*cm))
contenido.append(Paragraph("Detalle por Area", styles["Heading2"]))
contenido.append(Spacer(1, 0.3*cm))

datos_areas = [["Area","Este mes","Proximo","2 meses","OK","Total"]]
for area in areas:
    df_a = df[df["UBICACION"]==area]
    datos_areas.append([area,
        str(int((df_a["ESTADO"]=="🔴 ESTE MES").sum())),
        str(int((df_a["ESTADO"]=="🟡 PRÓXIMO MES").sum())),
        str(int((df_a["ESTADO"]=="🟠 EN 2 MESES").sum())),
        str(int((df_a["ESTADO"]=="🟢 OK").sum())),
        str(len(df_a))])
t2 = Table(datos_areas, colWidths=[6*cm,2.5*cm,2.5*cm,2.5*cm,2.5*cm,2*cm])
t2.setStyle(TableStyle([
    ("BACKGROUND",(0,0),(-1,0),colors.HexColor("#1F3864")),
    ("TEXTCOLOR",(0,0),(-1,0),colors.white),
    ("FONTNAME",(0,0),(-1,0),"Helvetica-Bold"),
    ("ALIGN",(1,0),(-1,-1),"CENTER"),
    ("FONTSIZE",(0,0),(-1,-1),8),
    ("GRID",(0,0),(-1,-1),0.5,colors.grey),
    ("ROWBACKGROUNDS",(0,1),(-1,-1),[colors.white,colors.HexColor("#F2F2F2")]),
]))
contenido.append(t2)
doc.build(contenido)
print("✅ PDF generado")

# ============================================================
# ENVIAR EMAILS
# ============================================================
print("\nEnviando emails...")

def construir_tabla_html(equipos_df):
    """Construye tabla HTML con los equipos del mes"""
    filas = ""
    for _, row in equipos_df.iterrows():
        filas += f"""
        <tr>
            <td style="padding:8px;border:1px solid #ddd;">{row['N_INVENTARIO']}</td>
            <td style="padding:8px;border:1px solid #ddd;">{row['DESCRIPCION']}</td>
            <td style="padding:8px;border:1px solid #ddd;">{row['UBICACION']}</td>
            <td style="padding:8px;border:1px solid #ddd;">{row['PERIODICIDAD']}</td>
        </tr>"""
    return f"""
    <table style="border-collapse:collapse;width:100%;font-size:13px;">
        <tr style="background:#1F3864;color:white;">
            <th style="padding:10px;border:1px solid #ddd;">N° Inventario</th>
            <th style="padding:10px;border:1px solid #ddd;">Equipo</th>
            <th style="padding:10px;border:1px solid #ddd;">Área</th>
            <th style="padding:10px;border:1px solid #ddd;">Periodicidad</th>
        </tr>
        {filas}
    </table>"""

def enviar_email(destinatario, equipos_df):
    """Envía email con equipos del mes al destinatario"""

    nombre = destinatario["nombre"]
    email  = destinatario["email"]
    areas_asignadas = destinatario["areas"]

    # Filtrar por áreas si tiene asignadas
    if areas_asignadas:
        df_dest = equipos_df[equipos_df["UBICACION"].isin(areas_asignadas)]
    else:
        df_dest = equipos_df

    total = len(df_dest)
    mes_nombre = NOMBRE_MES[MES_ACTUAL]

    tabla_html = construir_tabla_html(df_dest)

    cuerpo_html = f"""
    <html>
    <body style="font-family:Arial,sans-serif;max-width:800px;margin:0 auto;padding:20px;">

        <div style="background:#1F3864;color:white;padding:20px;border-radius:8px 8px 0 0;">
            <h2 style="margin:0;">🏥 ESE Hospital San Rafael Yolombó</h2>
            <p style="margin:5px 0 0 0;">Sistema de Gestión de Mantenimiento Biomédico</p>
        </div>

        <div style="background:#f9f9f9;padding:20px;border:1px solid #ddd;">
            <p>Hola <strong>{nombre}</strong>,</p>

            <p>Este es el reporte automático de mantenimiento para el mes de 
            <strong>{mes_nombre} {HOY.year}</strong>.</p>

            <div style="background:#FF4444;color:white;padding:15px;border-radius:6px;margin:15px 0;">
                <h3 style="margin:0;">⚠️ {total} equipos requieren mantenimiento este mes</h3>
            </div>

            <h3>Equipos a revisar en {mes_nombre}:</h3>
            {tabla_html}

            <br>
            <p>Se adjuntan los reportes completos en Excel y PDF.</p>

            <div style="background:#1F3864;color:white;padding:15px;border-radius:6px;margin:15px 0;">
                <p style="margin:0;">📋 <strong>Recuerda:</strong> Marca cada equipo como revisado 
                una vez completes el mantenimiento.</p>
            </div>
        </div>

        <div style="background:#f0f0f0;padding:10px;text-align:center;font-size:11px;color:#666;">
            Generado automáticamente — {HOY.strftime('%d/%m/%Y %H:%M')} |
            Ingeniería Biomédica — ESE Hospital San Rafael Yolombó
        </div>

    </body>
    </html>"""

    # Crear mensaje
    msg = MIMEMultipart("alternative")
    msg["Subject"] = f"🏥 Mantenimiento {mes_nombre} {HOY.year} — {total} equipos pendientes"
    msg["From"]    = EMAIL_REMITENTE
    msg["To"]      = email

    msg.attach(MIMEText(cuerpo_html, "html"))

    # Adjuntar Excel
    for archivo in [ARCHIVO_EXCEL, ARCHIVO_PDF]:
        if os.path.exists(archivo):
            with open(archivo, "rb") as f:
                parte = MIMEBase("application", "octet-stream")
                parte.set_payload(f.read())
                encoders.encode_base64(parte)
                parte.add_header("Content-Disposition", f"attachment; filename={os.path.basename(archivo)}")
                msg.attach(parte)

    # Enviar
    try:
        with smtplib.SMTP_SSL("smtp.gmail.com", 465) as servidor:
            servidor.login(EMAIL_REMITENTE, PASSWORD_APP)
            servidor.sendmail(EMAIL_REMITENTE, email, msg.as_string())
        print(f"  ✅ Email enviado a {nombre} ({email}) — {total} equipos")
    except Exception as e:
        print(f"  ❌ Error enviando a {email}: {e}")

# Enviar a cada destinatario
columnas_email = ["ITEM","DESCRIPCION","UBICACION","N_INVENTARIO","PERIODICIDAD","ESTADO"]
equipos_mes = df[df["ESTADO"]=="🔴 ESTE MES"][columnas_email]

for destinatario in DESTINATARIOS:
    enviar_email(destinatario, equipos_mes)

print("\n" + "="*60)
print("✅ V4 COMPLETA")
print(f"Emails enviados: {len(DESTINATARIOS)}")
print(f"Equipos este mes: {len(equipos_mes)}")
print("="*60)