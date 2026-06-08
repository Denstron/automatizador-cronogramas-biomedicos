"""
AUTOMATIZADOR DE CRONOGRAMAS BIOMÉDICOS V3
==========================================
- Detecta áreas automáticamente
- Genera reporte Excel con filtros por área
- Genera PDF resumen general
"""

import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm

ARCHIVO_ENTRADA = "Cronograma_de_mantenimiento_ESE_Hospital_Yolombo.xlsx"
ARCHIVO_EXCEL   = "Reporte_Cronograma_Equipos_Biomedicos.xlsx"
ARCHIVO_PDF     = "Reporte_Cronograma_Equipos_Biomedicos.pdf"

HOY = datetime.today()
MES_ACTUAL = HOY.month

MESES = {
    "ENERO":1,"FEBRERO":2,"MARZO":3,"ABRIL":4,"MAYO":5,"JUNIO":6,
    "JULIO":7,"AGOSTO":8,"SEPTIEMBRE":9,"OCTUBRE":10,"NOVIEMBRE":11,"DICIEMBRE":12
}
MESES_INVERSO = {v: k for k, v in MESES.items()}

COLORES_EXCEL = {
    "🔴 ESTE MES":    "FF4444",
    "🔴 VENCIDO":     "FF4444",
    "🟡 PRÓXIMO MES": "FFD700",
    "🟠 EN 2 MESES":  "FFA500",
    "🟢 OK":          "44BB44",
    "SIN FECHA":      "CCCCCC",
}
COLOR_ENCABEZADO = "1F3864"
COLOR_FILA_PAR   = "F2F2F2"

# ============================================================
# PROCESAR DATOS
# ============================================================
print("="*60)
print("AUTOMATIZADOR DE CRONOGRAMAS BIOMÉDICOS V3")
print("="*60)
print(f"Fecha: {HOY.strftime('%d/%m/%Y')}\n")

df = pd.read_excel(ARCHIVO_ENTRADA, header=0)
df.columns = ["ITEM","DESCRIPCION","UBICACION","N_INVENTARIO","FABRICANTE","MARCA",
              "MODELO","SERIE","CLASIFICACION_RIESGO","REGISTRO_INVIMA",
              "FECHA_COMPRA","PERIODICIDAD","FECHA_MANTENIMIENTO"]
df = df.dropna(subset=["DESCRIPCION"])
df.replace("NR","",inplace=True)

def extraer_meses(texto):
    if pd.isna(texto) or str(texto).strip()=="": return []
    return [MESES[p.strip()] for p in str(texto).upper().split("/") if p.strip() in MESES]

def calcular_estado(meses):
    if not meses: return "SIN FECHA", None
    meses = sorted(meses)
    proximo = next((m for m in meses if m >= MES_ACTUAL), None)
    if proximo is None: return "🔴 VENCIDO", meses[-1]
    diff = proximo - MES_ACTUAL
    if diff == 0: return "🔴 ESTE MES", proximo
    elif diff == 1: return "🟡 PRÓXIMO MES", proximo
    elif diff <= 2: return "🟠 EN 2 MESES", proximo
    else: return "🟢 OK", proximo

df["MESES_PARSED"] = df["FECHA_MANTENIMIENTO"].apply(extraer_meses)
resultados = df["MESES_PARSED"].apply(calcular_estado)
df["ESTADO"] = resultados.apply(lambda x: x[0])
df["PROXIMO_MES"] = resultados.apply(lambda x: MESES_INVERSO.get(x[1],"") if x[1] else "")

# Detectar áreas automáticamente
areas = sorted(df["UBICACION"].dropna().unique().tolist())
print(f"Áreas detectadas: {len(areas)}")
for area in areas:
    print(f"  - {area}")

columnas = ["ITEM","DESCRIPCION","UBICACION","N_INVENTARIO",
            "PERIODICIDAD","FECHA_MANTENIMIENTO","PROXIMO_MES","ESTADO"]
df_reporte = df[columnas].copy()

# ============================================================
# GENERAR EXCEL CON HOJA POR ÁREA
# ============================================================
print(f"\nGenerando Excel: {ARCHIVO_EXCEL}...")

with pd.ExcelWriter(ARCHIVO_EXCEL, engine="openpyxl") as writer:
    # Hoja con todos los equipos
    df_reporte.to_excel(writer, sheet_name="TODOS", index=False)

    # Hoja por cada área
    for area in areas:
        df_area = df_reporte[df_reporte["UBICACION"] == area]
        nombre_hoja = area[:31]  # Excel limita a 31 caracteres
        df_area.to_excel(writer, sheet_name=nombre_hoja, index=False)

# Aplicar colores a todas las hojas
wb = load_workbook(ARCHIVO_EXCEL)

def aplicar_estilos(ws):
    for i, ancho in enumerate([6,35,25,15,12,22,15,18], 1):
        ws.column_dimensions[ws.cell(1,i).column_letter].width = ancho
    for cell in ws[1]:
        cell.fill = PatternFill("solid", fgColor=COLOR_ENCABEZADO)
        cell.font = Font(bold=True, color="FFFFFF", size=10)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 25
    borde = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )
    for idx, fila in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=2):
        estado = ws.cell(idx, 8).value
        color_fila = COLOR_FILA_PAR if idx % 2 == 0 else "FFFFFF"
        for cell in fila[:-1]:
            cell.fill = PatternFill("solid", fgColor=color_fila)
            cell.alignment = Alignment(vertical="center")
            cell.border = borde
        color_estado = COLORES_EXCEL.get(estado, "FFFFFF")
        fila[-1].fill = PatternFill("solid", fgColor=color_estado)
        fila[-1].font = Font(bold=True, color="FFFFFF" if estado in ["🔴 ESTE MES","🔴 VENCIDO","🟢 OK"] else "000000")
        fila[-1].alignment = Alignment(horizontal="center", vertical="center")
        fila[-1].border = borde
    ws.freeze_panes = "A2"

for hoja in wb.sheetnames:
    aplicar_estilos(wb[hoja])

# Hoja RESUMEN
ws_r = wb.create_sheet("RESUMEN", 0)
ws_r.column_dimensions["A"].width = 30
ws_r.column_dimensions["B"].width = 12
ws_r.column_dimensions["C"].width = 12
ws_r.column_dimensions["D"].width = 12
ws_r.column_dimensions["E"].width = 12
ws_r.column_dimensions["F"].width = 12

encabezados = ["ÁREA", "🔴 ESTE MES", "🟡 PRÓXIMO", "🟠 2 MESES", "🟢 OK", "TOTAL"]
ws_r.append(encabezados)

for cell in ws_r[1]:
    cell.fill = PatternFill("solid", fgColor=COLOR_ENCABEZADO)
    cell.font = Font(bold=True, color="FFFFFF", size=10)
    cell.alignment = Alignment(horizontal="center")

for area in areas:
    df_a = df[df["UBICACION"] == area]
    fila = [
        area,
        int((df_a["ESTADO"]=="🔴 ESTE MES").sum()),
        int((df_a["ESTADO"]=="🟡 PRÓXIMO MES").sum()),
        int((df_a["ESTADO"]=="🟠 EN 2 MESES").sum()),
        int((df_a["ESTADO"]=="🟢 OK").sum()),
        len(df_a)
    ]
    ws_r.append(fila)

# Fila TOTAL
ws_r.append([
    "TOTAL HOSPITAL",
    int((df["ESTADO"]=="🔴 ESTE MES").sum()),
    int((df["ESTADO"]=="🟡 PRÓXIMO MES").sum()),
    int((df["ESTADO"]=="🟠 EN 2 MESES").sum()),
    int((df["ESTADO"]=="🟢 OK").sum()),
    len(df)
])
ultima = ws_r.max_row
for cell in ws_r[ultima]:
    cell.font = Font(bold=True)
    cell.fill = PatternFill("solid", fgColor="DDDDDD")

wb.save(ARCHIVO_EXCEL)
print(f"✅ Excel guardado: {ARCHIVO_EXCEL}")

# ============================================================
# GENERAR PDF RESUMEN
# ============================================================
print(f"Generando PDF: {ARCHIVO_PDF}...")

doc = SimpleDocTemplate(ARCHIVO_PDF, pagesize=A4,
    topMargin=2*cm, bottomMargin=2*cm,
    leftMargin=2*cm, rightMargin=2*cm)

styles = getSampleStyleSheet()
titulo_style = ParagraphStyle("titulo", parent=styles["Title"],
    fontSize=16, textColor=colors.HexColor("#1F3864"), spaceAfter=8)
sub_style = ParagraphStyle("sub", parent=styles["Normal"],
    fontSize=10, textColor=colors.grey, spaceAfter=20)

contenido = []

# Título
contenido.append(Paragraph("Reporte de Cronograma de Equipos Biomedicos", titulo_style))
contenido.append(Paragraph(f"Fecha: {HOY.strftime('%d de %B de %Y')} | Total equipos: {len(df)}", sub_style))

# Tabla resumen general
contenido.append(Paragraph("Resumen General", styles["Heading2"]))
contenido.append(Spacer(1, 0.3*cm))

datos_resumen = [
    ["Estado", "Equipos", "%"],
    ["🔴 Este mes", str(int((df["ESTADO"]=="🔴 ESTE MES").sum())),
     f"{int((df['ESTADO']=='🔴 ESTE MES').sum())/len(df)*100:.1f}%"],
    ["🟡 Próximo mes", str(int((df["ESTADO"]=="🟡 PRÓXIMO MES").sum())),
     f"{int((df['ESTADO']=='🟡 PRÓXIMO MES').sum())/len(df)*100:.1f}%"],
    ["🟠 En 2 meses", str(int((df["ESTADO"]=="🟠 EN 2 MESES").sum())),
     f"{int((df['ESTADO']=='🟠 EN 2 MESES').sum())/len(df)*100:.1f}%"],
    ["🟢 OK", str(int((df["ESTADO"]=="🟢 OK").sum())),
     f"{int((df['ESTADO']=='🟢 OK').sum())/len(df)*100:.1f}%"],
    ["TOTAL", str(len(df)), "100%"],
]

tabla_resumen = Table(datos_resumen, colWidths=[8*cm, 4*cm, 4*cm])
tabla_resumen.setStyle(TableStyle([
    ("BACKGROUND", (0,0), (-1,0), colors.HexColor("#1F3864")),
    ("TEXTCOLOR", (0,0), (-1,0), colors.white),
    ("FONTNAME", (0,0), (-1,0), "Helvetica-Bold"),
    ("ALIGN", (0,0), (-1,-1), "CENTER"),
    ("BACKGROUND", (0,1), (-1,1), colors.HexColor("#FF4444")),
    ("TEXTCOLOR", (0,1), (-1,1), colors.white),
    ("BACKGROUND", (0,2), (-1,2), colors.HexColor("#FFD700")),
    ("BACKGROUND", (0,3), (-1,3), colors.HexColor("#FFA500")),
    ("BACKGROUND", (0,4), (-1,4), colors.HexColor("#44BB44")),
    ("TEXTCOLOR", (0,4), (-1,4), colors.white),
    ("FONTNAME", (0,-1), (-1,-1), "Helvetica-Bold"),
    ("GRID", (0,0), (-1,-1), 0.5, colors.grey),
    ("ROWBACKGROUNDS", (0,5), (-1,5), [colors.HexColor("#DDDDDD")]),
]))
contenido.append(tabla_resumen)
contenido.append(Spacer(1, 0.8*cm))

# Tabla por área
contenido.append(Paragraph("Detalle por Área", styles["Heading2"]))
contenido.append(Spacer(1, 0.3*cm))

datos_areas = [["Área", "🔴 Este mes", "🟡 Próximo", "🟠 2 meses", "🟢 OK", "Total"]]
for area in areas:
    df_a = df[df["UBICACION"] == area]
    datos_areas.append([
        area,
        str(int((df_a["ESTADO"]=="🔴 ESTE MES").sum())),
        str(int((df_a["ESTADO"]=="🟡 PRÓXIMO MES").sum())),
        str(int((df_a["ESTADO"]=="🟠 EN 2 MESES").sum())),
        str(int((df_a["ESTADO"]=="🟢 OK").sum())),
        str(len(df_a))
    ])

tabla_areas = Table(datos_areas, colWidths=[6*cm, 2.5*cm, 2.5*cm, 2.5*cm, 2.5*cm, 2*cm])
tabla_areas.setStyle(TableStyle([
    ("BACKGROUND", (0,0), (-1,0), colors.HexColor("#1F3864")),
    ("TEXTCOLOR", (0,0), (-1,0), colors.white),
    ("FONTNAME", (0,0), (-1,0), "Helvetica-Bold"),
    ("ALIGN", (1,0), (-1,-1), "CENTER"),
    ("ALIGN", (0,0), (0,-1), "LEFT"),
    ("FONTSIZE", (0,0), (-1,-1), 8),
    ("GRID", (0,0), (-1,-1), 0.5, colors.grey),
    ("ROWBACKGROUNDS", (0,1), (-1,-1), [colors.white, colors.HexColor("#F2F2F2")]),
]))
contenido.append(tabla_areas)

doc.build(contenido)
print(f"✅ PDF guardado: {ARCHIVO_PDF}")
print("\n✅ V3 COMPLETA")
print(f"Áreas procesadas: {len(areas)}")
